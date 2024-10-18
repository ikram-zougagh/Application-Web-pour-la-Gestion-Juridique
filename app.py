from io import BytesIO
from datetime import datetime
import openpyxl
from flask import Flask, render_template,request,session,send_file,redirect,url_for
import mysql.connector
import re
from openpyxl.utils import get_column_letter
app=Flask(__name__)
app.secret_key = 'cairocoders-ednalan'
DB_HOST = "localhost"
DB_NAME = "sys"
DB_USER = "root"  # Replace with your MySQL username
DB_PASS = "barae123"  # Replace with your MySQL password
conn = mysql.connector.connect(
    host=DB_HOST,
    database=DB_NAME,
    user=DB_USER,
    password=DB_PASS
)
def check_login(username, password):
    # Vérifier que le nom d'utilisateur est "admin" et le mot de passe est "123456"
    if username == "admin" and password == "123456&&":
        session['logged_in'] = True
        return True
    else:
        return False
@app.route('/log',methods=['GET','POST'])
def log():
#vérifier si les champs de login son remplit
 if request.method == 'POST' and 'username' in request.form and 'password' in request.form:
       username = request.form['username']
       password = request.form['password']
       print(username)
       print(password)
       # Vérifier les informations de login
       if check_login(username, password):
           # Ajouter le nom d'utilisateur à la session pour le garder connecté
             session['username'] = username
             #return redirect(url_for('dashboard'))
             return render_template('acceuil.html', username=session['username'])
       else:
           # Afficher un message d'erreur si les informations de connexion sont incorrectes
           error_message = "إسم المستخدم أو الرقم السري غير صحيح"
           return render_template('log_in.html', error_message=error_message)
 return render_template('log_in.html')
@app.route('/suivie.html',methods=['GET','POST'])
def suivie():
    # Le code qui permet de suivre les dossiers
    results = None
    message = None

    if request.method == "POST":
        search_term = request.form.get("search_term")

        # Vérifier si la variable search_term correspond au format "nombre/nombre/nombre"
        if re.match(r'^\d+/\d+/\d+$', search_term):
            cursor = conn.cursor()
            query = "SELECT * FROM dossiers JOIN etats ON dossiers.etat_idEtat = etats.idEtat WHERE dossiers.numDossier = %s"
            cursor.execute(query, (search_term,))
            records = cursor.fetchall()
            print(records)

            # Remplacer les valeurs vides par une valeur par défaut (par exemple, une chaîne vide ou un message)
            def replace_empty_value(value):
                return value if value else ""  # Vous pouvez remplacer "Valeur vide" par ce que vous voulez afficher à la place

            results = []
            for record in records:
                # Appliquer la fonction de remplacement pour chaque élément de l'enregistrement
                modified_record = [replace_empty_value(value) for value in record]
                results.append(modified_record)

            if not results:
                message = "نتائج البحث فارغة"
        else:
            message = "رقم الملف غير صحيح"

    return render_template('suivie.html', results=results, message=message)
@app.route('/ajout-dossier.html',methods=['GET','POST'])
def ajout_dossier():
    #code qui permet d'ajouter des dossier
    if request.method == 'POST':
        numero = request.form.get('numero')
        type_str = request.form.get('type')
        type_int = int(re.search(r'\d+', type_str).group()) if type_str else None
        print(type_int)
        annee = request.form.get('annee')
        date_liquidation = request.form.get('date_liquidation') # Set default value to None
        date_notification = request.form.get('date_notification')  # Set default value to None
        date_mise_en_oeuvre = request.form.get('date_mise_en_oeuvre')  # Set default value to None
        date_preservation = request.form.get('date_preservation')  # Set default value to None
        date_action_pour_delit = request.form.get('date_action_pour_delit')  # Set default value to None
        date_deliquant = request.form.get('date_deliquant')  # Set default value to None
        remarques = request.form.get('remarques')
        numero_dossier=annee+"/"+str(type_int)+"/"+numero
        date_liquidation = date_liquidation if date_liquidation else None
        date_notification= date_notification if  date_notification else None
        date_mise_en_oeuvre = date_mise_en_oeuvre  if date_mise_en_oeuvre  else None
        date_preservation = date_preservation if date_preservation else None
        date_action_pour_delit = date_action_pour_delit if date_action_pour_delit else None
        date_deliquant= date_deliquant if date_deliquant else None

        # Insérer les données dans la table Etats
        cursor = conn.cursor()
        insert_etats_query = "INSERT INTO Etats (notification, liquidation, mise_en_oueuvre, deliquant, preservation, action_pour_delit) VALUES (%s, %s, %s, %s, %s, %s)"
        etats_data = (date_notification, date_liquidation, date_mise_en_oeuvre,date_deliquant, date_preservation, date_action_pour_delit)
        cursor.execute(insert_etats_query, etats_data)
        conn.commit()

        # Récupérer l'ID de l'enregistrement inséré dans la table Etats
        etat_id = cursor.lastrowid
        # Insérer les données dans la table dossiers en utilisant l'ID de l'état correspondant
        insert_dossiers_query = "INSERT INTO dossiers (numDossier, remarque, etat_idEtat) VALUES (%s, %s, %s)"
        dossiers_data = (numero_dossier, remarques, etat_id)
        cursor.execute(insert_dossiers_query, dossiers_data)
        conn.commit()
        cursor.close()
    return render_template('ajout-dossier.html')
# Fonction pour extraire les enregistrements de la base de données entre les IDs de début et de fin
def get_records(start_id, end_id):
    try:
        cursor = conn.cursor()
        query = "SELECT id,numDossier,remarque,notification,liquidation, mise_en_oueuvre, deliquant, preservation, action_pour_delit  FROM dossiers JOIN etats ON dossiers.etat_idEtat = etats.idEtat WHERE id BETWEEN %s AND %s"
        cursor.execute(query,(start_id,end_id))
        records = cursor.fetchall()
        cursor.close()
        conn.close()
        print(records)
        return records

    except Exception as e:
        print(f"Erreur lors de la récupération des enregistrements: {e}")
        return []
@app.route('/extraction-donnees.html', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        start_id = request.form['start_id']
        end_id = request.form['end_id']
        records = get_records(start_id, end_id)

        # Créez un fichier Excel avec les enregistrements
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Entêtes
        #sheet.append(['ملاحظة', 'الجنحي', 'العقوبات الحبسية', 'الحفظ', 'التنفيذ', 'التبليغ',
                 #'التصفية', 'رقم الملف'])
        headers = ['معرف','رقم الملف','ملاحظة', 'التبليغ','التصفية','التنفيذ','الحفظ','العقوبات الحبسية','الجنحي']
        sheet.append(headers)
        # Ajouter les enregistrements
        for record in records:
            record_data = [str(value) if isinstance(value, datetime) else value for value in record]
            sheet.append(record_data)
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            adjusted_width = (length + 2) * 1.2
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width
        # Créez un flux de bytes pour stocker le fichier Excel
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        current_date = datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
        filename = f'بيانات مستخرجة_{current_date}.xlsx'
        return send_file(output, attachment_filename=filename,as_attachment=True)

    return render_template('extraction-donnees.html')
@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return render_template('log_in.html')
@app.route('/autre.html', methods=['GET', 'POST'])
def autre_nombre():
    results = None
    message = None
    if request.method == "POST":
        search_term = request.form.get("search_term")
        cursor = conn.cursor()

        # Exécutez la requête pour obtenir le nombre total de dossiers et le nombre de dossiers dans chaque état
        query = """
                SELECT
                    COUNT(*) AS total_dossiers,
                    SUM(CASE WHEN etats.notification IS NOT NULL THEN 1 ELSE 0 END) AS nombre_notification,
                    SUM(CASE WHEN etats.liquidation IS NOT NULL THEN 1 ELSE 0 END) AS nombre_liquidation,
                    SUM(CASE WHEN etats.mise_en_oueuvre IS NOT NULL THEN 1 ELSE 0 END) AS nombre_mise_en_oueuvre,
                    SUM(CASE WHEN etats.deliquant IS NOT NULL THEN 1 ELSE 0 END) AS nombre_deliquant,
                    SUM(CASE WHEN etats.preservation IS NOT NULL THEN 1 ELSE 0 END) AS nombre_preservation,
                    SUM(CASE WHEN etats.action_pour_delit IS NOT NULL THEN 1 ELSE 0 END) AS nombre_action_pour_delit
                FROM
                    dossiers
                LEFT JOIN
                    etats
                ON
                    dossiers.etat_idEtat = etats.idEtat
            """
        cursor.execute(query)
        results = cursor.fetchall()

    if not results:
        message = ""
    return render_template('autre.html', results=results ,message=message)


@app.route('/autre_supp', methods=['POST'])
def autre_supp():
    if request.method == 'POST':
        numero_dossier = request.form.get('numero_dossier')

        # Connectez-vous à la base de données
        cursor = conn.cursor()

        # Vérifiez si le dossier avec le numéro de dossier donné existe
        query = "SELECT * FROM dossiers WHERE numDossier = %s"
        cursor.execute(query, (numero_dossier,))
        dossier = cursor.fetchone()

        if not dossier:
            # Si le dossier n'existe pas, renvoyez un message d'erreur
            message = "رقم الملف الذي أدخلته غير موجود"
            return render_template('autre.html', message=message)

        # Supprimez le dossier de la base de données
        delete_query = "DELETE d, e FROM dossiers AS d LEFT JOIN etats AS e ON d.etat_idEtat = e.idEtat WHERE d.numDossier = %s;"
        cursor.execute(delete_query, (numero_dossier,))
        conn.commit()

        # Fermez la connexion à la base de données
        cursor.close()
        conn.close()

        # Redirigez l'utilisateur vers la page "autre.html" avec un message de succès
        message = f" تم حذف ملف رقم {numero_dossier}بنجاح"
        return render_template('autre.html', message=message)

@app.route('/acceuil')
def home():
        return render_template('acceuil.html')
if __name__=="__main__":
    app.run(debug=True)